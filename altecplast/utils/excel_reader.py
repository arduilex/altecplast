"""
Module pour lire les données Excel et alimenter le dashboard
"""

from django.conf import settings
from openpyxl import load_workbook
from altecplast.utils import CELLS_PROD, CELLS_PESAGE, CELLS_STOCK
from altecplast.utils import print_error
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import zipfile

def get_dashboard_data(request_date):
    """
    Lit le fichier Excel et retourne les KPIs du dashboard

    Args:
        month (str): Numéro du mois (01-12). Si None, utilise le mois d'août par défaut.
    """
    try:
        data = {}
        today_date = datetime.today().strftime("%Y-%m")
        # Trouver le fichier Excel pour le mois demandé
        year, month = request_date.split('-')
        excel_pesage_file = find_excel_pesage_file(year, month)
        wb_pesage = open_wb(excel_pesage_file)

        # recap mensuel du mois en production
        data["production"] = read_prod_mensuel(wb_pesage)

        # Lire les données de pesages des feuilles
        data["fournisseur"] = group_reading(wb_pesage, 'mp', ["kg", "mad"])
        data["client"] = group_reading(wb_pesage, 'pf', ["kg", "mad"])
        sort_group_by_priority(data["fournisseur"])
        sort_group_by_priority(data["client"])

        # si on n'est pas au mois actuel, changer de fichier, on veut celui d'aujourd'hui !
        if request_date != today_date:
            year, month = today_date.split('-')
            excel_pesage_file = find_excel_pesage_file(year, month)
            wb_pesage.close()
            wb_pesage = open_wb(excel_pesage_file)
        excel_stock_file  = find_excel_stock_file(year)
        wb_stock = open_wb(excel_stock_file)
        data["today"] = {}
        data["today"]["pesage"] = read_daily_pesage(wb_pesage)
        data["today"]["prod"] = read_daily_prod(wb_pesage)
        data["today"]["stock"] = read_daily_stock(wb_stock, month)

        # Fermer les classeurs
        wb_pesage.close()
        wb_stock.close()

        # Retourner les données
        return data

    except Exception as e:
        print_error("Erreur get_dashboard_data", e)
        return None

 # Petite fonction utilitaire pour formater les nombres
 
def debug_result(result):
    import json
    print(json.dumps(result, indent=2, ensure_ascii=False))

def open_wb(file_dir):
    if not file_dir:
        print_error("Fichier Excel non trouvé")
        return None

    # Valider que le fichier est bien un ZIP valide (format Excel .xlsx)
    if not zipfile.is_zipfile(file_dir):
        print_error(f"Le fichier n'est pas un fichier Excel valide (pas un ZIP): {file_dir}")
        return None

    # Charger le classeur Excel
    try:
        wb = load_workbook(str(file_dir), read_only=True, data_only=True)
    except Exception as e:
        print_error(f"Erreur lors de l'ouverture du fichier Excel: {file_dir}", e)
        return None
    return wb

def find_excel_stock_file(year):
    """
    Trouve le fichier Excel STOCK correspondant à l'année demandé

    Args:
        year (str) : Année AAAA

    Returns:
        Path: Chemin vers le fichier Excel ou None si non trouvé
    """
    try:
        # Racine du projet
        project_root = Path(settings.ROOT_DIR)
        # Collecter tous les fichiers Excel valides
        # Chercher le dossier qui commence par le numéro du mois
        for folder_root in project_root.iterdir():
            if folder_root.is_dir() and "DOSSIER" in folder_root.name and year in folder_root.name:
                folder_stock = folder_root / "DOSSIER PRODUCTION"
                for file in folder_stock.iterdir():
                    if file.name.startswith("FICHE DE STOCK"):
                        return file

        print_error(f"Aucun fichier Excel FICHE DE STOCK {year} trouvé")
        return None
    
    except Exception as e:
        print_error(f"Erreur lors de la recherche du fichier Excel FICHE DE STOCK", e)
        return None

def find_excel_pesage_file(year, month):
    """
    Trouve le fichier Excel pesage correspondant au mois et à l'année demandé

    Args:
        year (str) : Année AAAA
        month (str): Numéro du mois mm (01-12)

    Returns:
        Path: Chemin vers le fichier Excel ou None si non trouvé
    """
    try:
        # Racine du projet
        project_root = Path(settings.ROOT_DIR)
        # Collecter tous les fichiers Excel valides
        # Chercher le dossier qui commence par le numéro du mois
        for folder_root in project_root.iterdir():
            if folder_root.is_dir() and "DOSSIER" in folder_root.name and year in folder_root.name:
                folder_pesage = folder_root / "PESAGE ENTREE SORTIE"
                for folder_month in folder_pesage.iterdir():
                    if folder_month.is_dir() and folder_month.name.startswith(f"{month}_"):
                        for file in folder_month.iterdir():
                            if file.is_file() and file.suffix.lower() in ['.xlsx', '.xlsm'] and not file.name.startswith('~$'):
                                return file

        print_error(f"Aucun fichier Excel PESAGE trouvé pour le mois {month}")
        return None

    except Exception as e:
        print_error(f"Erreur lors de la recherche du fichier Excel PESAGE", e)
        return None

def sort_group_by_priority(data):
    """
    Trie les v_type_pes de chaque dict dans data avec la priorité suivante :
    1. Clés sans le mot clé "retour"
    2. Clés avec le mot clé "retour"
    """
    # Parcourir chaque label (kg, mad, etc.)
    for label in data:
        # Récupérer les clés v_type_pes
        type_pes_keys = list(data[label].keys())

        # Séparer les clés sans et avec "retour"
        keys_without_retour = []
        keys_with_retour = []

        for key in type_pes_keys:
            if key and "retour" in key.lower():
                keys_with_retour.append(key)
            else:
                keys_without_retour.append(key)

        # Trier chaque groupe alphabétiquement
        keys_without_retour.sort()
        keys_with_retour.sort()

        # Reconstruire le dictionnaire dans le nouvel ordre
        sorted_dict = {}
        for key in keys_without_retour:
            sorted_dict[key] = data[label][key]
        for key in keys_with_retour:
            sorted_dict[key] = data[label][key]

        # Remplacer le dictionnaire par la version triée
        data[label] = sorted_dict

def group_reading(wb, sheet, request_label):
    try:
        CONFIG = CELLS_PESAGE[sheet]
        ws_name = CONFIG["sheet_name"]
        # Vérifier si la feuille existe
        if ws_name not in wb.sheetnames:
            print_error("Feuille {ws_name} non trouvée")
            return {}
        #chargement des paramètres de lectures 
        ws = wb[ws_name]
        LABELS = CONFIG["labels"]
        k_type_pes = LABELS["type_pes"]
        k_type_prod = LABELS["type_prod"]
        k_type_forme = LABELS["forme"]
        data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

        for row_data in ws.iter_rows(min_row=3, max_col=20, values_only=True):
            if not row_data[0]:
                break # si la ligne est vide, on sort de la boucle
            v_type_pes   = row_data[k_type_pes]
            v_type_prod  = row_data[k_type_prod]
            v_type_forme = row_data[k_type_forme]
            v_categorie = f"{v_type_prod}_{v_type_forme}"
            # fusionner entrée reférencée et non référencée
            if v_type_pes in ["Entrée référencée", "Entrée non référencée"]:
                v_type_pes = "Entrées"
            for label in request_label:
               k = LABELS[label]  # titre de la valeur
               v = row_data[k]   # valeur numérique
               if not isinstance(v, (int, float)):
                   v = 0   # Si la cellule ne comorte as de nombre, on enregistre 0
               data[label][v_type_pes][v_categorie] += abs(v)
        return data
    except Exception as e:
        print_error(f"Erreur group_reading dans {sheet}", e)
        return {}

def read_labels_values_rows(ws, row, col_s, col_e):
    try: 
        row_label = next(ws.iter_rows(min_col=col_s, max_col=col_e, min_row=CELLS_PROD["row_label"], max_row=CELLS_PROD["row_label"], values_only=True)) 
        row_data = next(ws.iter_rows(min_col=col_s, max_col=col_e, min_row=row, max_row=row, values_only=True))
        return {k:v for k, v in zip(row_label, row_data)}
    except Exception as e:
        print_error(f"Erreur read_labels_values_rows", e)
        return {}

def read_mp(ws, cf):
    data =  read_labels_values_rows(ws, cf["row"], cf["col_start"], cf["col_end"]-1)
    return {k:v for k, v in data.items() if v}

def read_prod_mensuel(wb):
    try: 
        def read_cells(ws, items):
            try:
                result = {}
                for item in items:
                    value = ws[item["cell"]].value
                    if value:
                        result[item["id"]] = f"{round(float(value), 2):,} {item['unit']}".replace(",", " ").replace(".", ",")
                    else:
                        result[item["id"]] = ''
                return result

            except Exception as e:
                print_error(f"Erreur dans read_prod_mensuel", e)
                return {}
        
        def read_range(ws, items):
                values_dict = {}
                colors_list = []
                
                noms = items["nom"]
                valeurs = items["valeur"]
                colors = items["colors"]
                
                # Créer une liste de tuples (nom, valeur, couleur) pour faciliter le tri
                data_tuples = []
                for cell_name, cell_value, color in zip(ws[noms], ws[valeurs], colors):
                    value = cell_value[0].value
                    if isinstance(value, int):
                        if value > 0:
                            data_tuples.append((cell_name[0].value, value, color))
                    else:
                        break
                
                # Reconstruire les dictionnaires/listes triés
                for name, value, color in data_tuples:
                    values_dict[name] = value
                    colors_list.append(color)
                return {
                    "data": values_dict,
                    "colors": colors_list
                }

        
        def read_prealpina_prod(ws, cf):
            try:
                row = cf["row"]
                col_start, col_end = cf["cols_range_jour"]
                prod_jour = read_labels_values_rows(ws, row, col_start, col_end)
                col_start, col_end = cf["cols_range_nuit"]
                prod_nuit = read_labels_values_rows(ws, row, col_start, col_end)
                data = {}
                colors = []
                for (k_j, v_j), (k_n, v_n), c in zip(prod_jour.items(), prod_nuit.items(), cf["colors"]):
                    if v_j:
                        colors.append(c)
                        data[k_j] = v_j
                        if k_n in prod_jour:
                            data[k_n] += v_n
                        else:
                            data[k_n] = v_n
                return {
                    "data": data,
                    "colors": colors
                }
            except Exception as e:
                print_error(f"Erreur read_prod_mensuel", e)
                return {
                    "data": {},
                    "colors": []
                }

        result = {}
        ws = wb["Recap mensuel"]

        result["lavage"] = read_cells(ws, CELLS_PROD["lavage"])
        result["prealpina"] = read_cells(ws, CELLS_PROD["prealpina"])
        result["lavage"]["diagram"] = read_range(ws, CELLS_PROD["lavage_type"])

        ws = wb["Base de donnée P°"]
        result["prealpina"]["diagram"] = read_prealpina_prod(ws, CELLS_PROD["prealpina_type"])
        result["mp_souple"]   = read_mp(ws, CELLS_PROD["mp_souple"])
        result["mp_rigide"]   = read_mp(ws, CELLS_PROD["mp_rigide"])
        result["mp_broye"]    = read_mp(ws, CELLS_PROD["mp_broye"])
        result["mp_densifie"] = read_mp(ws, CELLS_PROD["mp_densifie"])

        return result
    

    except Exception as e:
        print_error(f"Erreur read_prod_mensuel", e)
        return {
            "data": {},
            "colors": []
        }

def read_daily_pesage(wb):
    try:
        # Trouver le fichier DailyData.csv
        day_pesage = {}

        daily_info_value = CELLS_PESAGE["daily_info_value"]  # Indices des colonnes à garder
        daily_info_unit = CELLS_PESAGE["daily_info_unit"]  # untier de chaque valeur 

        ws = wb["Feuille de pesage"]
        for row_data in ws.iter_rows(min_row=3, values_only=True):
            if not row_data[0]:
                break
            date = row_data[1].date()
            if date == datetime.today().date():
                type_pesage = row_data[0]

                # Extraire les données selon les colonnes spécifiées
                data = []
                for col, unit in zip(daily_info_value, daily_info_unit):
                    if not row_data[col]:
                        data.append("-")
                    else:
                        data.append(f"{row_data[col]} {unit}")

                # Créer la clé du dictionnaire si elle n'existe pas
                if type_pesage not in day_pesage:
                    day_pesage[type_pesage] = []

                # Ajouter la ligne de données au type correspondant
                day_pesage[type_pesage].append(data)
                day_pesage["date"] = date.strftime("%d/%m/%Y")  # on enregistre la date pour afficher la date réel dans l'onglet aujourd'hui du dashboard

        return day_pesage

    except Exception as e:
        print_error(f"Erreur read_daily_pesage", e)
        return {}

def read_daily_prod(wb):
    
    try:
        day_prod = {}
        ws = wb["Base de donnée P°"]
        last_row = ws.max_row
        prod_data = [cell.value for cell in ws[last_row]]
        day_prod = {
                "date": prod_data[0].strftime("%d/%m/%Y"),
                "lav": {
                    "prod_total": prod_data[5],
                    "jour": {
                        "prod_jour": prod_data[6],
                        "noir_c": prod_data[7],
                        "noir_nc": prod_data[8],
                        "blanc_c": prod_data[9],
                        "blanc_nc": prod_data[10],
                        "bleu_c": prod_data[11],
                        "bleu_nc": prod_data[12],
                        "heure_w": prod_data[13],
                        "nombre_h": prod_data[14],
                        "arret1": prod_data[15],
                        "arret1_h": prod_data[16],
                        "arret2": prod_data[17],
                        "arret2_h": prod_data[18],
                        "motif1": prod_data[19],
                        "motif2": prod_data[20],
                    },
                    "nuit": {
                        "prod_nuit": prod_data[21],
                        "noir_c": prod_data[22],
                        "noir_nc": prod_data[23],
                        "blanc_c": prod_data[24],
                        "blanc_nc": prod_data[25],
                        "bleu_c": prod_data[26],
                        "bleu_nc": prod_data[27],
                        "heure_w": prod_data[28],
                        "nombre_h": prod_data[29],
                        "arret1": prod_data[30],
                        "arret1_h": prod_data[31],
                        "arret2": prod_data[32],
                        "arret2_h": prod_data[33],
                        "motif1": prod_data[34],
                        "motif2": prod_data[35],
                    },
                    "nc": prod_data[36],
                    "purge_j": prod_data[37],
                    "purge_n": prod_data[38],
                },
                "alp": {
                    "prod_total": prod_data[39],
                    "jour": {
                        "prod_jour": prod_data[40],
                        "bb_noir": prod_data[41],
                        "bb_blanc": prod_data[42],
                        "sacs_blanc": prod_data[43],
                        "sacs_bleu": prod_data[44],
                        "sacs_noir": prod_data[45],
                        "autre": prod_data[46],
                        "heure_w": prod_data[47],
                        "nombre_h": prod_data[48],
                        "arret1": prod_data[49],
                        "arret1_h": prod_data[50],
                        "arret2": prod_data[51],
                        "arret2_h": prod_data[52],
                        "motif1": prod_data[53],
                        "motif2": prod_data[54],
                    },
                    "nuit": {
                        "prod_nuit": prod_data[55],
                        "bb_noir": prod_data[56],
                        "bb_blanc": prod_data[57],
                        "sacs_blanc": prod_data[58],
                        "sacs_bleu": prod_data[59],
                        "sacs_noir": prod_data[60],
                        "autre": prod_data[61],
                        "heure_w": prod_data[62],
                        "nombre_h": prod_data[63],
                        "arret1": prod_data[64],
                        "arret1_h": prod_data[65],
                        "arret2": prod_data[66],
                        "arret2_h": prod_data[67],
                        "motif1": prod_data[68],
                        "motif2": prod_data[69],
                    },
                    "nc": prod_data[70],
                    "purge_j": prod_data[71],
                    "purge_n": prod_data[72],
                },
                "mp": {
                    "mp_souple": read_labels_values_rows(ws, last_row, CELLS_PROD["mp_souple"]["col_start"], CELLS_PROD["mp_souple"]["col_end"]),
                    "mp_rigide": read_labels_values_rows(ws, last_row, CELLS_PROD["mp_rigide"]["col_start"], CELLS_PROD["mp_rigide"]["col_end"]),
                    "mp_broye": read_labels_values_rows(ws, last_row, CELLS_PROD["mp_broye"]["col_start"], CELLS_PROD["mp_broye"]["col_end"]),
                    "mp_densifie": read_labels_values_rows(ws, last_row, CELLS_PROD["mp_densifie"]["col_start"], CELLS_PROD["mp_densifie"]["col_end"]),
                }
            }
        return day_prod

    except Exception as e:
            print_error("Erreur read_daily_prod", e)
            return {}

def read_daily_stock(wb, month):
    try:
        stock_day = {}
        month_int = int(month)
        sheeet_name = datetime(2000, month_int, 1).strftime('%B').upper()
        ws = wb[sheeet_name]
        densifie_keys = CELLS_STOCK["densifie"].keys()
        granule_keys  = CELLS_STOCK["granule"].keys()
        next_index_key = len(densifie_keys) + 1
        for stock_data in ws.iter_rows(min_row=4, values_only=True):
            if stock_data[1] is not None:
                stock_day = {
                    "date": stock_data[0].strftime("%d/%m/%Y"),
                    "stock_densifie": {k: stock_data[i + 1] or 0 for i, k in enumerate(densifie_keys)},
                    "stock_granule":  {k: stock_data[i + next_index_key] or 0 for i, k in enumerate(granule_keys)},
                    
                }
        stock_day["colors"] = CELLS_STOCK
        return stock_day
                
    except Exception as e:
        print_error(f"Erreur lors de la lecture du fichier Excel STOCK", e)
        return None

if __name__ == "__main__":
    data = get_dashboard_data('01-2025')
    print(data)